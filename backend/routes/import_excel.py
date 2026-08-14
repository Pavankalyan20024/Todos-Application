from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import require_roles
from ..models import AuditLog, Member, User

router = APIRouter(tags=["excel"])
MAX_UPLOAD = 5 * 1024 * 1024

ALIASES = {
    "sno": "employee_id", "serial no": "employee_id", "employee id": "employee_id", "id": "employee_id",
    "team member": "team_member", "member name": "team_member", "employee name": "team_member",
    "project/workstream": "project_workstream", "project workstream": "project_workstream", "project": "project_workstream", "workstream": "project_workstream",
    "role": "role", "access level": "access_level", "access role": "access_level", "system role": "access_level", "permission level": "access_level",
    "current work": "current_work", "technologies used": "technologies", "technology": "technologies", "technologies": "technologies", "technology/skills": "technologies",
    "skills/expertise": "skills", "skills": "skills", "working with": "working_with", "contribution/value add": "contribution_value_add", "contribution": "contribution_value_add",
    "status": "status", "expected completion date": "expected_completion_date", "expected completion": "expected_completion_date", "completion date": "expected_completion_date",
    "blockers/support required": "blockers", "blockers": "blockers", "support required": "blockers",
}
REQUIRED = {"employee_id", "team_member", "project_workstream", "role", "current_work", "status", "expected_completion_date"}
STATUSES = {"ongoing": "Ongoing", "completed": "Completed", "complete": "Completed", "on hold": "On Hold", "not started": "Not Started", "blocked": "Blocked"}


def clean_header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").replace(" - ", " ").replace(" / ", "/").split()).strip(".:")


def values(value):
    return [item.strip() for item in str(value or "").split(",") if item.strip() and item.strip().lower() not in {"none", "n/a", "no blockers"}]


@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(require_roles("admin"))):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please select a supported .xlsx Excel file.")
    content = await file.read(MAX_UPLOAD + 1)
    if len(content) > MAX_UPLOAD:
        raise HTTPException(413, "Excel file must be 5 MB or smaller.")
    try:
        sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
    except Exception:
        raise HTTPException(400, "Unable to read the Excel workbook.")
    rows = list(sheet.iter_rows(values_only=True))
    header_index = mapping = None
    for index, row in enumerate(rows[:20]):
        candidate = {column: ALIASES.get(clean_header(value)) for column, value in enumerate(row)}
        candidate = {column: field for column, field in candidate.items() if field}
        if mapping is None or len(candidate) > len(mapping):
            header_index, mapping = index, candidate
    if not mapping or not REQUIRED.issubset(set(mapping.values())):
        raise HTTPException(422, "Excel file is missing one or more required columns.")
    imported = updated = skipped = 0
    errors = []
    for row_number, row in enumerate(rows[header_index + 1:], header_index + 2):
        raw = {field: row[column] if column < len(row) else None for column, field in mapping.items()}
        if not any(value not in (None, "") for value in raw.values()):
            continue
        try:
            missing = [field for field in REQUIRED if raw.get(field) in (None, "")]
            if missing:
                raise ValueError("missing required values")
            employee_id = str(raw["employee_id"]).strip()
            member = db.query(Member).filter(Member.employee_id == employee_id).first()
            existed = bool(member)
            member = member or Member(employee_id=employee_id)
            member.team_member = str(raw["team_member"]).strip()
            member.project_workstream = str(raw["project_workstream"]).strip()
            member.role = str(raw["role"]).strip()
            if raw.get("access_level") not in (None, ""):
                access_level = str(raw["access_level"]).strip().lower()
                if access_level not in {"admin", "manager", "employee"}:
                    raise ValueError("invalid access level")
                member.access_level = access_level
            elif not existed:
                member.access_level = "employee"
            member.current_work = str(raw["current_work"]).strip()
            member.working_with = str(raw.get("working_with") or "").strip()
            member.contribution_value_add = str(raw.get("contribution_value_add") or "").strip()
            member.technologies = values(raw.get("technologies"))
            member.skills = values(raw.get("skills"))
            member.blockers = values(raw.get("blockers"))
            member.status = STATUSES.get(str(raw["status"]).strip().lower(), str(raw["status"]).strip())
            if member.status not in set(STATUSES.values()):
                raise ValueError("invalid status")
            date_value = raw["expected_completion_date"]
            member.expected_completion_date = date_value.date() if isinstance(date_value, datetime) else date_value if isinstance(date_value, date) else datetime.fromisoformat(str(date_value)).date()
            db.add(member)
            db.flush()
            updated += int(existed)
            imported += int(not existed)
        except Exception as error:
            skipped += 1
            errors.append({"row": row_number, "message": str(error)})
    db.add(AuditLog(user_id=user.id, role=user.role, action="excel_imported", entity_type="member", entity_id=file.filename))
    db.commit()
    return {"success": True, "imported": imported, "updated": updated, "skipped": skipped, "errors": errors[:50]}
